'''
Postcondition 1: Program takes all relevant data (facility name, patient name, visits, minutes, status codes) from all therapy records in chosen directory and places it in the Therapy Dashboard.xlsm
in a newly created sheet named with the facility, therapy discipline, and month
'''
from tkinter import *   ## notice lowercase 't' in tkinter here

import os
import subprocess
import openpyxl
import tkinter
from tkinter import filedialog
import time

# excel_document = openpyxl.load_workbook(filename)

def main():
    Tk().withdraw # we don't want a full GUI, so keep the root window from appearing

    filedirectory = filedialog.askdirectory(title="Open Folder",
     initialdir=('G:\\Therapy Charting Grids\\'))

    therapy_dashboard = openpyxl.load_workbook('G:\Therapy Record Interfaces\Front Range Therapy\Dashboard\Therapy Dashboard.xlsm', read_only=False, keep_vba=True)

    therapy_sheet = therapy_dashboard.sheetnames
    therapy_sheet = therapy_sheet[0]
    therapy_sheet = therapy_dashboard[therapy_sheet]
    
    os.chdir(filedirectory)

    discipline = "Null"
    
    # Subgoal 1: Here we are defining month/discipline/facility variables from the chosen file path stored in the variable filedirectory
    if " January " in filedirectory:
        month = "January"
    elif " February " in filedirectory:
        month = "February"
    elif " March " in filedirectory:
        month = "March"
    elif " April " in filedirectory:
        month = "April"
    elif " May " in filedirectory:
        month = "May"
    elif " June " in filedirectory:
        month = "June"
    elif " July " in filedirectory:
        month = "July"
    elif " August " in filedirectory:
        month = "August"
    elif " September " in filedirectory:
        month = "September"
    elif " October " in filedirectory:
        month = "October"
    elif " November " in filedirectory:
        month = "November"
    elif " December " in filedirectory:
        month = "December"
    if (r"/OT") in filedirectory:
        discipline = "OT"
    elif (r"/PT") in filedirectory:
        discipline = "PT"
    elif (r"/ST") in filedirectory:
        discipline = "ST"
    if "Centre Avenue" in filedirectory:
        facility = "Centre Avenue"
    elif "Columbine Commons" in filedirectory:
        facility = "Columbine Commons"
    elif (r"North Shore") in filedirectory:
        facility = "North Shore"
    elif "Lemay Avenue" in filedirectory:
        facility = "Lemay Avenue"
    elif "Columbine West" in filedirectory:
        facility = "Columbine West"
    # Subgoal 2: Create new worksheet from template(sheet1) rename according to month/discipline/facility variables and use worksheet to to hold all values retrieved from therapy records
    therapy_dashboard.copy_worksheet(therapy_sheet)
    therapy_sheet = therapy_dashboard.sheetnames
    therapy_sheet = therapy_sheet[-1]
    therapy_sheet = therapy_dashboard[therapy_sheet]
    therapy_sheet.title = month + " " + facility + " " + discipline
    therapy_sheet = month + " " + facility + " " + discipline
    # Postcondition 1: All relevant data is retrieved from therapy records and placed in worksheet
    for filename in os.listdir(filedirectory):
        if ".xlsm" in filename and "~" not in filename:
            print("Processing", filename)
            excel_document = openpyxl.load_workbook(filename=filename, read_only=True, data_only=True)
            sheet = excel_document.get_sheet_names()[0]
            patient_first_name = excel_document[sheet]["O3"].value
            patient_last_name = excel_document[sheet]["H3"].value
            patient_name = "{}, {}".format(patient_last_name, patient_first_name)
            named_ranges = excel_document.get_named_ranges()
            visits_dict = {}
            facility_name = excel_document[sheet]["W2"].value
            for named_range in named_ranges:
                address_split = named_range.attr_text.split("!")
                address = address_split[1]
                if named_range.name == "TotalMinutes":
                    total_minutes = excel_document[sheet][address].value
                if named_range.name == "PreviousMonthVisits":
                    previous_month_visits = excel_document[sheet][address].value
                if named_range.name == "CurrentMonthVisits":
                    current_month_visits = excel_document[sheet][address].value
                cell_dict = {"B": 6, "C": 7, "D": 8, "E": 9, "F": 10, "G": 11, "H": 12, "I": 13, "J": 14, "K": 15, "L": 16, "M": 17, "N": 18,
                             "O": 19, "P": 20, "Q": 21, "R": 22,"S": 23, "T": 24, "U": 25, "V": 26, "W": 27, "X": 28, "Y": 29, "Z": 30, "AA": 31,
                             "AB": 32, "AC": 33, "AD": 34, "AE": 35, "AF": 36}
                for cell in cell_dict.keys():
                    if named_range.name == "{}{}{}".format("Initials", cell, "1"):
                        if excel_document[sheet][address].value is None:
                            pass
                        else:
                            visits_dict["{}{}{}".format("Initials", cell, "1")] = "X"
                            break
                    if named_range.name == "{}{}".format("TreatmentMinutes", cell):
                        if excel_document[sheet][address].value == "R":
                            visits_dict["{}{}".format("TreatmentMinutes", cell)] = "R"
                        elif excel_document[sheet][address].value == "A":
                            visits_dict["{}{}".format("TreatmentMinutes", cell)] = "A"
                        elif excel_document[sheet][address].value == "B":
                            visits_dict["{}{}".format("TreatmentMinutes", cell)] = "B"
                        elif excel_document[sheet][address].value == "C":
                            visits_dict["{}{}".format("TreatmentMinutes", cell)] = "C"
                        elif excel_document[sheet][address].value == "D":
                            visits_dict["{}{}".format("TreatmentMinutes", cell)] = "D"
            try:
                total_visits = previous_month_visits + current_month_visits
            except TypeError:
                total_visits = current_month_visits
            for num in range(1, 1000):
                    if therapy_dashboard[therapy_sheet].cell(row=num, column=1).value is None:
                        therapy_dashboard[therapy_sheet].cell(row=num, column=1).value = filename
                        therapy_dashboard[therapy_sheet].cell(row=num, column=2).value = patient_name
                        therapy_dashboard[therapy_sheet].cell(row=num, column=3).value = discipline
                        therapy_dashboard[therapy_sheet].cell(row=num, column=4).value = total_visits
                        therapy_dashboard[therapy_sheet].cell(row=num, column=5).value = total_minutes
                        for cell, x in visits_dict.items():
                            for cell_2, num_2 in cell_dict.items():
                                if cell == "{}{}{}".format("Initials", cell_2, "1"):
                                    therapy_dashboard[therapy_sheet].cell(row=num, column=num_2).value = x
                        for cell, x in visits_dict.items():
                            for cell_2, num_2 in cell_dict.items():
                                if cell == "{}{}".format("TreatmentMinutes", cell_2):
                                    therapy_dashboard[therapy_sheet].cell(row=num, column=num_2).value = x
                        break
    
    home_dir = os.path.normpath("G:/Therapy Record Interfaces/Front Range Therapy/Dashboard/")
    os.chdir(home_dir)
    therapy_dashboard.save('Therapy Dashboard.xlsm')

    os.startfile('G:\Therapy Record Interfaces\Front Range Therapy\Dashboard\Therapy Dashboard.xlsm')

main()
